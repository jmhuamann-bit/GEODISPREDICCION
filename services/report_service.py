"""Generacion de reportes descargables: panel ejecutivo en PDF (reportlab) y el Centro de
Monitoreo en Excel (openpyxl). El export a CSV vive en app/api/monitoring.py (mas liviano)."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from services import dashboard_service, monitoring_service

NAVY = colors.HexColor("#0A1F44")
TEAL = colors.HexColor("#00A89A")
GRAY = colors.HexColor("#64748B")
LIGHT = colors.HexColor("#F1F5F9")


def build_executive_pdf() -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=1.6 * cm, bottomMargin=1.6 * cm, leftMargin=1.8 * cm, rightMargin=1.8 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("GDTitle", parent=styles["Title"], textColor=NAVY, fontSize=20, spaceAfter=2)
    subtitle_style = ParagraphStyle("GDSubtitle", parent=styles["Normal"], textColor=GRAY, fontSize=10.5, spaceAfter=16)
    h2_style = ParagraphStyle("GDH2", parent=styles["Heading2"], textColor=NAVY, fontSize=13, spaceBefore=14, spaceAfter=8)
    body_style = ParagraphStyle("GDBody", parent=styles["Normal"], fontSize=9.5, textColor=colors.HexColor("#334155"))

    kpis = dashboard_service.get_kpis()
    corredores = dashboard_service.get_top_corredores_riesgo(limit=10)

    elements = []
    elements.append(Paragraph("GEODIS Decision Intelligence Platform", title_style))
    elements.append(Paragraph(
        f"Reporte Ejecutivo — generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style
    ))

    elements.append(Paragraph("Indicadores Ejecutivos", h2_style))
    kpi_rows = [
        ["Embarques totales", f"{kpis['total_embarques']:,}", "OTIF", f"{kpis['otif_pct']}%"],
        ["Lead Time promedio", f"{kpis['lead_time_promedio_dias']} días", "Nivel de servicio", f"{kpis['nivel_servicio_pct']}%"],
        ["Costo total", f"$ {kpis['costo_total_cop']:,.0f} COP", "Costo promedio", f"$ {kpis['costo_promedio_cop']:,.0f} COP"],
        ["CO₂ estimado", f"{kpis['co2_toneladas']:,.1f} ton", "Contingencias activas", f"{kpis['contingencias_activas']:,}"],
        ["Riesgo Alto/Crítico", f"{kpis['distribucion_riesgo']['alto_critico']:,} embarques",
         "Riesgo Bajo", f"{kpis['distribucion_riesgo']['bajo']:,} embarques"],
    ]
    kpi_table = Table(kpi_rows, colWidths=[4.3 * cm, 4 * cm, 4.3 * cm, 4 * cm])
    kpi_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TEXTCOLOR", (0, 0), (0, -1), GRAY), ("TEXTCOLOR", (2, 0), (2, -1), GRAY),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"), ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (1, 0), (1, -1), NAVY), ("TEXTCOLOR", (3, 0), (3, -1), NAVY),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(kpi_table)

    elements.append(Paragraph("Corredores Logísticos con Mayor Riesgo", h2_style))
    corredor_rows = [["Corredor", "Embarques", "OTIF", "Riesgo Promedio"]]
    for c in corredores:
        corredor_rows.append([c["corredor"], f"{c['total_embarques']:,}", f"{c['otif_pct']}%", f"{c['riesgo_promedio_pct']}%"])
    corredor_table = Table(corredor_rows, colWidths=[6.5 * cm, 3.2 * cm, 3.2 * cm, 3.7 * cm])
    corredor_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(corredor_table)

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Reporte generado automáticamente por GEODIS Decision Intelligence Platform. "
        "Los niveles de riesgo provienen del modelo de Predicciones IA vigente al momento de la generación.",
        body_style,
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def build_monitoring_excel(filtros: dict) -> bytes:
    result = monitoring_service.search_shipments(filtros, page=1, page_size=5000)
    items = result["items"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoreo GEODIS"

    headers = ["ID Viaje", "Fecha", "Sector Cliente", "Origen", "Destino", "Corredor", "Transporte",
               "Prioridad", "Lead Time (d)", "OTIF", "Costo Total (COP)", "Riesgo (%)", "Nivel Riesgo"]
    header_fill = PatternFill(start_color="0A1F44", end_color="0A1F44", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item["id_viaje"])
        ws.cell(row=row_idx, column=2, value=item["fecha"])
        ws.cell(row=row_idx, column=3, value=item["sector_cliente"])
        ws.cell(row=row_idx, column=4, value=item["municipio_origen"])
        ws.cell(row=row_idx, column=5, value=item["municipio_destino"])
        ws.cell(row=row_idx, column=6, value=item["corredor_logistico"])
        ws.cell(row=row_idx, column=7, value=item["tipo_transporte"])
        ws.cell(row=row_idx, column=8, value=item["prioridad_cliente"])
        ws.cell(row=row_idx, column=9, value=item["leadtime_real_dias"])
        ws.cell(row=row_idx, column=10, value="Cumplido" if item["otif"] else "Incumplido")
        ws.cell(row=row_idx, column=11, value=item["costo_total_cop"])
        ws.cell(row=row_idx, column=12, value=item["prob_riesgo_incumplimiento"])
        ws.cell(row=row_idx, column=13, value=item["nivel_riesgo"])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
